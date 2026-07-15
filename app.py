"""
Promotion Agreement Comparator — Web App
=========================================
Streamlit version: drag & drop files in the browser, get a report back.
Run locally with:  streamlit run app.py
"""

import io
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook

TOLERANCE = 0.05  # RM 0.05 rounding tolerance
DASH_VALUES = ['—', '-', '–', 'nan', 'None', '', ' ', '\xa0']


def is_empty(val):
    if val is None:
        return True
    if pd.isna(val) if not isinstance(val, str) else False:
        return True
    return str(val).strip() in DASH_VALUES


def read_supplier_reply(file):
    df_raw = pd.read_excel(file, header=None, engine='openpyxl')
    records = []
    for i, row in df_raw.iterrows():
        if i < 8:
            continue
        plu = row.iloc[0]
        if is_empty(plu) or str(plu).strip() == '0':
            continue
        try:
            plu = int(float(str(plu).strip()))
        except:
            continue
        if plu == 0:
            continue

        def safe(idx):
            v = row.iloc[idx] if idx < len(row) else None
            return None if is_empty(v) else v

        records.append({
            "PLU":                  plu,
            "Product":              str(row.iloc[1]).strip(),
            "Mechanic":             str(row.iloc[2]).strip() if not is_empty(row.iloc[2]) else None,
            "Normal_Cost_ExclGST":  safe(4),
            "Normal_RSP_InclGST":   safe(7),
            "Promo_RSP_InclGST":    safe(11),
            "Promo_Rebate_ExclGST": safe(12),
            "Member_Point":         safe(13),
            "Member_Rebate":        safe(14),
        })
    return records


def read_watson_agreement(file):
    df_raw = pd.read_excel(file, header=None, engine='openpyxl')

    header_row = None
    for i, row in df_raw.iterrows():
        vals = [str(v).upper() for v in row.values if pd.notna(v)]
        if any('PLU' in v for v in vals):
            header_row = i
            break

    if header_row is None:
        st.error("Could not find a header row (containing 'PLU') in the Watson Agreement file.")
        st.stop()

    if hasattr(file, "seek"):
        file.seek(0)  # rewind if it's an in-memory/uploaded file object

    # Header genuinely spans 2 rows: a group label (e.g. "Cost (RM)") on the
    # first row, and a sub-label (e.g. "Normal" / "Promo") on the second.
    # Data starts right after, on header_row + 2.
    df = pd.read_excel(file, header=[header_row, header_row + 1], engine='openpyxl')

    new_cols = []
    for col in df.columns:
        parts = []
        items = col if isinstance(col, tuple) else [col]
        for c in items:
            s = str(c).strip().replace('\xa0', ' ').replace('\u2019', "'")
            if s not in ['nan', ''] and 'Unnamed' not in s:
                parts.append(s)
        new_cols.append(' | '.join(parts) if parts else 'unknown')
    df.columns = new_cols

    records = []
    for _, row in df.iterrows():
        plu = None
        for col in df.columns:
            if 'PLU CODE' in col.upper() or col.upper().startswith('PLU'):
                try:
                    v = str(row[col]).strip().replace('\xa0', '')
                    plu = int(float(v))
                    if plu > 0:
                        break
                except:
                    pass
        if not plu or plu == 0:
            continue

        def get_col(*keywords):
            kw_upper = [k.upper() for k in keywords]
            for col in df.columns:
                col_upper = col.upper().replace('\xa0', ' ')
                if all(k in col_upper for k in kw_upper):
                    v = row[col]
                    return None if is_empty(v) else v
            return None

        mech = get_col('PROMOTION MECHANICS', 'PROMO 1') or get_col('PROMO', '1')
        if is_empty(mech):
            mech = None

        records.append({
            "PLU":           plu,
            "Mechanic":      str(mech).strip() if mech else None,
            "Cost_Normal":   get_col('COST', 'NORMAL'),
            "Cost_Promo":    get_col('COST', 'PROMO'),
            "Rebate_Promo1": get_col('REBATE', 'PROMO 1'),
            "Rebate_Member": get_col('REBATE', 'MEMBER'),
            "Price_Normal":  get_col('PRICE', 'NORMAL'),
            "Price_Promo":   get_col('PRICE', 'PROMO'),
            "Price_Member":  get_col('PRICE', 'MEMBER'),
        })
    return records


def num_ok(a, b):
    if is_empty(a) and is_empty(b):
        return True
    if is_empty(a) or is_empty(b):
        return False
    try:
        return abs(float(str(a).replace(',', '')) - float(str(b).replace(',', ''))) <= TOLERANCE
    except (ValueError, TypeError):
        return False


def norm_mech(s):
    if is_empty(s):
        return None
    s = str(s).upper().strip()
    s = s.replace('\xa0', '').replace(' ', '').replace('AT', '@').replace('ND@', 'ND@')
    s = s.replace('2NDAT', '2ND@').replace('2ND AT', '2ND@')
    return s


def compare(supplier, watson):
    watson_map = {r['PLU']: r for r in watson}
    supplier_map = {r['PLU']: r for r in supplier}

    issues = []
    ok_count = 0

    for plu, s in supplier_map.items():
        if plu not in watson_map:
            issues.append({
                "PLU": plu, "Product": s['Product'], "Field": "PLU",
                "Issue": "❌ In Supplier Reply but NOT in Watson Agreement",
                "Supplier Value": plu, "Watson Value": "—"
            })
            continue

        w = watson_map[plu]
        row_issues = []

        sm = norm_mech(s.get('Mechanic'))
        wm = norm_mech(w.get('Mechanic'))
        if sm and wm and sm != wm:
            if sm not in wm and wm not in sm:
                row_issues.append({
                    "PLU": plu, "Product": s['Product'], "Field": "Mechanic",
                    "Issue": "⚠️ Mismatch",
                    "Supplier Value": s.get('Mechanic'), "Watson Value": w.get('Mechanic')
                })

        num_checks = [
            ("Normal Cost (Excl GST)",  s.get('Normal_Cost_ExclGST'), w.get('Cost_Normal')),
            ("Normal RSP (Incl GST)",   s.get('Normal_RSP_InclGST'),  w.get('Price_Normal')),
            ("Promo RSP (Incl GST)",    s.get('Promo_RSP_InclGST'),   w.get('Price_Promo')),
            ("Promo Rebate (Excl GST)", s.get('Promo_Rebate_ExclGST'), w.get('Rebate_Promo1')),
            ("Member Rebate",           s.get('Member_Rebate'),        w.get('Rebate_Member')),
        ]

        for field, sv, wv in num_checks:
            if is_empty(sv) and is_empty(wv):
                continue
            if not num_ok(sv, wv):
                row_issues.append({
                    "PLU": plu, "Product": s['Product'], "Field": field,
                    "Issue": "⚠️ Mismatch",
                    "Supplier Value": sv if not is_empty(sv) else "—",
                    "Watson Value":   wv if not is_empty(wv) else "—"
                })

        if row_issues:
            issues.extend(row_issues)
        else:
            ok_count += 1

    for plu in watson_map:
        if plu not in supplier_map:
            issues.append({
                "PLU": plu, "Product": "—", "Field": "PLU",
                "Issue": "❌ In Watson Agreement but NOT in Supplier Reply",
                "Supplier Value": "—", "Watson Value": plu
            })

    return issues, ok_count


def build_report_bytes(issues):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Report"

    headers = ["PLU", "Product", "Field", "Issue", "Supplier Value", "Watson Value"]

    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    red_fill = PatternFill("solid", start_color="FFCCCC")
    yel_fill = PatternFill("solid", start_color="FFF2CC")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for r, issue in enumerate(issues, 2):
        fill = red_fill if "❌" in issue["Issue"] else yel_fill
        for c, key in enumerate(["PLU", "Product", "Field", "Issue", "Supplier Value", "Watson Value"], 1):
            cell = ws.cell(row=r, column=c, value=str(issue.get(key, "")))
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ───────────────────────── UI ─────────────────────────

st.set_page_config(page_title="Promotion Agreement Comparator", page_icon="◇", layout="wide")

ACCENT = "#00e5c7"

st.markdown(
    f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;700&display=swap');

    html, body, [class*="css"] {{
        font-family: 'JetBrains Mono', 'Courier New', monospace !important;
    }}

    .block-container {{
        padding-top: 2.5rem;
        max-width: 1100px;
    }}

    /* Title */
    h1 {{
        letter-spacing: 0.06em;
        text-transform: uppercase;
        font-weight: 700 !important;
        font-size: 1.6rem !important;
        color: {ACCENT} !important;
        border-bottom: 1px solid rgba(0,229,199,0.25);
        padding-bottom: 0.6rem;
    }}

    /* Caption / subtitle */
    [data-testid="stCaptionContainer"] {{
        letter-spacing: 0.03em;
        opacity: 0.65;
    }}

    /* Metric cards */
    div[data-testid="stMetric"] {{
        background: #0d131b;
        border: 1px solid rgba(0,229,199,0.25);
        border-radius: 4px;
        padding: 14px 16px;
    }}
    div[data-testid="stMetric"] label {{
        text-transform: uppercase;
        letter-spacing: 0.05em;
        font-size: 0.72rem !important;
        opacity: 0.6;
    }}
    div[data-testid="stMetricValue"] {{
        color: {ACCENT} !important;
    }}

    /* File uploader */
    [data-testid="stFileUploaderDropzone"] {{
        background: #0d131b !important;
        border: 1px dashed rgba(0,229,199,0.35) !important;
        border-radius: 4px !important;
    }}

    /* Buttons */
    .stButton button, .stDownloadButton button {{
        border: 1px solid {ACCENT} !important;
        border-radius: 3px !important;
        letter-spacing: 0.05em;
        text-transform: uppercase;
        font-size: 0.85rem !important;
        transition: box-shadow 0.15s ease;
    }}
    .stButton button:hover, .stDownloadButton button:hover {{
        box-shadow: 0 0 10px rgba(0,229,199,0.45);
    }}

    /* Tabs */
    button[data-baseweb="tab"] {{
        font-family: 'JetBrains Mono', monospace !important;
        letter-spacing: 0.03em;
    }}
    button[aria-selected="true"] {{
        color: {ACCENT} !important;
    }}
    div[data-baseweb="tab-highlight"] {{
        background-color: {ACCENT} !important;
    }}

    hr, [data-testid="stDivider"] {{
        border-color: rgba(0,229,199,0.15) !important;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("◇ Promotion Agreement Comparator")
st.caption("// drag in a supplier reply, compare against the watson agreement, flag discrepancies")

st.divider()

col1, col2 = st.columns(2)
with col1:
    supplier_file = st.file_uploader(
        "SUPPLIER REPLY EXCEL", type=["xlsx", "xls"], key="supplier",
        help="Required every time — this is the file that changes."
    )
with col2:
    watson_file = st.file_uploader(
        "WATSON AGREEMENT EXCEL", type=["xlsx", "xls"], key="watson",
        help="Optional — leave blank to use the bundled default file, if one exists."
    )

st.write("")
compare_clicked = st.button(
    "▶ Compare", type="primary", use_container_width=True, disabled=(supplier_file is None)
)

if compare_clicked:
    with st.spinner("Reading files and comparing..."):
        try:
            supplier = read_supplier_reply(supplier_file)

            if watson_file is not None:
                watson = read_watson_agreement(watson_file)
            else:
                try:
                    with open("Watson Agreement.xlsx", "rb") as f:
                        watson = read_watson_agreement(io.BytesIO(f.read()))
                except FileNotFoundError:
                    st.error(
                        "No Watson Agreement file was uploaded, and no default "
                        "'Watson Agreement.xlsx' was found. Please upload one."
                    )
                    st.stop()

            issues, ok_count = compare(supplier, watson)

        except Exception as e:
            st.error(f"Something went wrong while reading the files: {e}")
            st.stop()

    st.divider()

    total_plu = ok_count + len({i["PLU"] for i in issues}) if issues else ok_count
    mismatches = [i for i in issues if "Mismatch" in i["Issue"]]
    missing = [i for i in issues if "❌" in i["Issue"]]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("TOTAL COMPARED", total_plu)
    m2.metric("MATCHED", ok_count)
    m3.metric("MISMATCHES", len(mismatches))
    m4.metric("MISSING", len(missing))

    if not issues:
        st.success("✓ PERFECT MATCH — no discrepancies found")
    else:
        report_bytes = build_report_bytes(issues)
        st.download_button(
            label="⬇ DOWNLOAD REPORT (.xlsx)",
            data=report_bytes,
            file_name="comparison_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        st.write("")
        search = st.text_input("FILTER — plu or product name", placeholder="e.g. 98079 or Cebion")

        def apply_filter(rows):
            if not search:
                return rows
            q = search.strip().lower()
            return [r for r in rows if q in str(r.get("PLU", "")).lower() or q in str(r.get("Product", "")).lower()]

        tab1, tab2 = st.tabs([f"⚠ MISMATCHES ({len(mismatches)})", f"✕ MISSING PLUs ({len(missing)})"])

        with tab1:
            filtered = apply_filter(mismatches)
            if filtered:
                st.dataframe(pd.DataFrame(filtered), use_container_width=True, hide_index=True)
            else:
                st.caption("No mismatches" + (" matching that filter." if search else "."))

        with tab2:
            filtered = apply_filter(missing)
            if filtered:
                st.dataframe(pd.DataFrame(filtered), use_container_width=True, hide_index=True)
            else:
                st.caption("No missing PLUs" + (" matching that filter." if search else "."))
